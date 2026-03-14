"""
Relational XLSX Exporter
=========================
Replaces the flat DataExporter.export_xlsx() with a multi-sheet workbook
designed for data analysts.

Workbook structure
------------------
Sheet 1 — RECORDS
    One row per processed form.
    Columns: form_id, template_id, processed_at, all scalar fields
    (full_name, matric_number, department, faculty, session, semester,
     level_100..400, mode_fulltime, mode_parttime, total_units, date)
    student_signature is excluded by design.

Sheet 2 — COURSES
    One row per course registration entry.
    Columns: form_id (FK → RECORDS), row_number, sn, course_code,
             course_title, unit
    form_id is the foreign key linking back to RECORDS.

Sheet 3 — VALIDATION_LOG
    One row per field per form — shows extraction confidence,
    validation status, and whether human correction was applied.
    Useful for analysing HITL effectiveness across the batch.

Usage
-----
    from project.output.relational_exporter import RelationalXLSXExporter

    exporter = RelationalXLSXExporter()
    path = exporter.export(structured_output, "outputs/form_001.xlsx")

    # Or batch export:
    path = exporter.export_batch(list_of_structured_outputs, "outputs/batch.xlsx")
"""

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style constants ────────────────────────────────────────────────────────────
_H_FILL  = PatternFill("solid", fgColor="1A1A2E")
_H_FONT  = Font(bold=True, color="FFFFFF", size=10)
_A_FILL  = PatternFill("solid", fgColor="E8F5E9")  # accepted — green tint
_F_FILL  = PatternFill("solid", fgColor="FFF9C4")  # flagged  — yellow
_C_FILL  = PatternFill("solid", fgColor="E3F2FD")  # corrected — blue tint
_R_FILL  = PatternFill("solid", fgColor="FFEBEE")  # rejected  — red tint
_THIN    = Side(style="thin", color="DDDDDD")
_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER  = Alignment(horizontal="center", vertical="center")
_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Fields that go on the RECORDS sheet (scalar, one value per form)
_RECORD_FIELDS = [
    "full_name", "matric_number", "department", "faculty",
    "level_100", "level_200", "level_300", "level_400",
    "session", "semester",
    "mode_fulltime", "mode_parttime",
    "total_units", "date",
    # student_signature deliberately excluded
]

# Fields that make up one course row
_COURSE_FIELDS = ["sn", "course_code", "course_title", "unit"]
_N_COURSE_ROWS = 3


class RelationalXLSXExporter:
    """
    Exports one or many structured DAPE outputs into a relational
    multi-sheet Excel workbook.
    """

    # ── Single-form export ─────────────────────────────────────────────────────

    def export(self, structured: dict, output_path: str) -> str:
        return self.export_batch([structured], output_path)

    # ── Batch export ───────────────────────────────────────────────────────────

    def export_batch(
        self,
        structured_list: list[dict],
        output_path: str,
    ) -> str:
        """
        Export a list of structured outputs into one relational workbook.

        Parameters
        ----------
        structured_list : list of dicts from OutputStructurer.structure()
        output_path     : destination .xlsx file path

        Returns
        -------
        str — path to written file
        """
        wb = openpyxl.Workbook()

        ws_records = wb.active
        ws_records.title = "RECORDS"
        ws_courses  = wb.create_sheet("COURSES")
        ws_vallog   = wb.create_sheet("VALIDATION_LOG")

        self._build_records(ws_records, structured_list)
        self._build_courses(ws_courses, structured_list)
        self._build_vallog(ws_vallog,   structured_list)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    # ── RECORDS sheet ──────────────────────────────────────────────────────────

    def _build_records(self, ws, records: list[dict]) -> None:
        cols = ["form_id", "template_id", "processed_at"] + _RECORD_FIELDS
        self._write_header(ws, cols)
        ws.freeze_panes = "A2"

        col_widths = {
            "form_id": 12, "template_id": 22, "processed_at": 22,
            "full_name": 28, "matric_number": 18, "department": 22,
            "faculty": 22, "session": 14, "semester": 12,
            "total_units": 11, "date": 13,
        }
        for key, w in col_widths.items():
            if key in cols:
                ws.column_dimensions[
                    get_column_letter(cols.index(key) + 1)
                ].width = w
        for key in ["level_100","level_200","level_300","level_400",
                    "mode_fulltime","mode_parttime"]:
            if key in cols:
                ws.column_dimensions[
                    get_column_letter(cols.index(key) + 1)
                ].width = 11

        for row_idx, s in enumerate(records, start=2):
            data   = s.get("data", {})
            row    = [
                s.get("form_id", ""),
                s.get("template_id", ""),
                s.get("processed_at", ""),
            ] + [data.get(f, "") for f in _RECORD_FIELDS]
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = _BORDER
                cell.alignment = _CENTER if col_idx > 3 and cols[col_idx-1].startswith(("level_","mode_")) else _LEFT
                cell.font      = Font(size=9)
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F5F5F5")

    # ── COURSES sheet ──────────────────────────────────────────────────────────

    def _build_courses(self, ws, records: list[dict]) -> None:
        cols = ["form_id", "row_number"] + _COURSE_FIELDS
        self._write_header(ws, cols)
        ws.freeze_panes = "A2"

        col_widths = {
            "form_id": 12, "row_number": 10,
            "sn": 8, "course_code": 14,
            "course_title": 40, "unit": 8,
        }
        for key, w in col_widths.items():
            if key in cols:
                ws.column_dimensions[
                    get_column_letter(cols.index(key) + 1)
                ].width = w

        data_row = 2
        for s in records:
            form_id = s.get("form_id", "")
            data    = s.get("data", {})
            alt     = data_row % 2 == 0
            fill    = PatternFill("solid", fgColor="F5F5F5") if alt else None

            for r in range(1, _N_COURSE_ROWS + 1):
                row = [
                    form_id,
                    r,
                    data.get(f"sn_{r}",           ""),
                    data.get(f"course_code_{r}",   ""),
                    data.get(f"course_title_{r}",  ""),
                    data.get(f"unit_{r}",          ""),
                ]
                for col_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=data_row, column=col_idx, value=val)
                    cell.border    = _BORDER
                    cell.alignment = _CENTER if col_idx in (1, 2, 6) else _LEFT
                    cell.font      = Font(size=9)
                    if fill:
                        cell.fill = fill
                data_row += 1

    # ── VALIDATION_LOG sheet ───────────────────────────────────────────────────

    def _build_vallog(self, ws, records: list[dict]) -> None:
        cols = [
            "form_id", "field_id", "field_type",
            "extracted_value", "final_value",
            "confidence", "validation_status", "validation_reason",
            "needs_review", "corrected",
        ]
        self._write_header(ws, cols)
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 26
        ws.column_dimensions["I"].width = 11
        ws.column_dimensions["J"].width = 10

        _STATUS_FILL = {
            "accepted":          _A_FILL,
            "corrected":         _C_FILL,
            "low_confidence":    _F_FILL,
            "semantic_failure":  _F_FILL,
            "rejected":          _R_FILL,
        }

        data_row = 2
        for s in records:
            form_id = s.get("form_id", "")
            for field in s.get("fields", []):
                status = field.get("validation_status", "")
                row = [
                    form_id,
                    field.get("field_id", ""),
                    field.get("field_type", ""),
                    str(field.get("value", "")),
                    str(field.get("final_value", "")),
                    round(float(field.get("confidence", 0.0)), 3),
                    status,
                    field.get("validation_reason", ""),
                    field.get("needs_review", False),
                    field.get("corrected", False),
                ]
                row_fill = _STATUS_FILL.get(status)
                for col_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=data_row, column=col_idx, value=val)
                    cell.border    = _BORDER
                    cell.alignment = _LEFT
                    cell.font      = Font(size=9)
                    if row_fill:
                        cell.fill = row_fill
                data_row += 1

    # ── Shared header writer ───────────────────────────────────────────────────

    @staticmethod
    def _write_header(ws, cols: list[str]) -> None:
        for col_idx, label in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = _H_FONT
            cell.fill      = _H_FILL
            cell.alignment = _CENTER
            cell.border    = _BORDER
        ws.row_dimensions[1].height = 28
