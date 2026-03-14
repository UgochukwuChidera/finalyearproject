"""
Excel → JSON Ground Truth Converter
=====================================
Reads the filled-in ground_truth_entry.xlsx and writes one JSON file
per form into ground_truth/form_001.json ... form_045.json.

Run after filling in the Excel spreadsheet:
    python excel_to_json.py
    python excel_to_json.py --input ground_truth/ground_truth_entry.xlsx
"""

import argparse
import json
from pathlib import Path

import openpyxl


BOOL_TRUE  = {"true",  "yes", "1", "x", "✓", "TRUE",  "True"}
BOOL_FALSE = {"false", "no",  "0", "",  "FALSE", "False"}
CHECKBOX_FIELDS = {
    "level_100", "level_200", "level_300", "level_400",
    "mode_fulltime", "mode_parttime",
}
COURSE_FIELDS = ["sn", "course_code", "course_title", "unit"]


def _coerce_bool(val) -> bool | None:
    if val is None or str(val).strip() == "":
        return None
    return str(val).strip().lower() in BOOL_TRUE


def _coerce_text(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def convert(
    input_path:  str = "ground_truth/ground_truth_entry.xlsx",
    output_dir:  str = "ground_truth",
    template_id: str = "student_academic_record",
) -> list[str]:
    wb       = openpyxl.load_workbook(input_path, data_only=True)
    ws_forms = wb["FORMS"]
    ws_courses = wb["COURSES"]

    # ── Read FORMS sheet ──────────────────────────────────────────────────────
    # Row 3 = headers, data starts row 4
    headers = [
        ws_forms.cell(row=3, column=c).value
        for c in range(1, ws_forms.max_column + 1)
    ]
    headers = [str(h).strip() if h else None for h in headers]

    forms: dict[str, dict] = {}
    for row in ws_forms.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        form_id = str(row[0]).strip()
        if not form_id.startswith("form_"):
            continue

        fields = {}
        for col_idx, header in enumerate(headers):
            if header is None or header == "form_id":
                continue
            val = row[col_idx] if col_idx < len(row) else None
            if header in CHECKBOX_FIELDS:
                fields[header] = _coerce_bool(val)
            else:
                fields[header] = _coerce_text(val)

        forms[form_id] = fields

    # ── Read COURSES sheet ────────────────────────────────────────────────────
    # Row 2 = headers, data starts row 3
    c_headers = [
        ws_courses.cell(row=2, column=c).value
        for c in range(1, ws_courses.max_column + 1)
    ]
    c_headers = [str(h).strip() if h else None for h in c_headers]

    courses: dict[str, list] = {}
    for row in ws_courses.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        form_id    = str(row[0]).strip()
        row_number = row[1] if len(row) > 1 else None

        entry = {}
        for col_idx, header in enumerate(c_headers):
            if header in (None, "form_id", "row_number"):
                continue
            val = row[col_idx] if col_idx < len(row) else None
            entry[header] = _coerce_text(val)

        # Only include non-empty course rows
        if any(v for v in entry.values()):
            courses.setdefault(form_id, []).append({
                "row": int(row_number) if row_number else len(courses.get(form_id, [])) + 1,
                **entry,
            })

    # ── Build per-form JSON ground truth ──────────────────────────────────────
    out_dir   = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    written   = []

    for form_id, fields in forms.items():
        # Flatten course table into individual fields for evaluator compatibility
        # AND keep structured version for the nested export
        form_courses = courses.get(form_id, [])
        for r in range(1, 4):
            matching = next((c for c in form_courses if c.get("row") == r), {})
            fields[f"sn_{r}"]           = matching.get("sn")
            fields[f"course_code_{r}"]  = matching.get("course_code")
            fields[f"course_title_{r}"] = matching.get("course_title")
            fields[f"unit_{r}"]         = matching.get("unit")

        gt = {
            "form_id":     form_id,
            "template_id": template_id,
            "fields":      fields,
            # Structured courses kept separately for nested XLSX export
            "_courses":    form_courses,
        }

        out_path = out_dir / f"{form_id}.json"
        with out_path.open("w", encoding="utf-8") as fh:
            json.dump(gt, fh, indent=2, ensure_ascii=False)
        written.append(str(out_path))

    return written


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert ground truth Excel → JSON")
    parser.add_argument("--input",       default="ground_truth/ground_truth_entry.xlsx")
    parser.add_argument("--output-dir",  default="ground_truth")
    parser.add_argument("--template-id", default="student_academic_record")
    args = parser.parse_args()

    files = convert(args.input, args.output_dir, args.template_id)
    print(f"Written {len(files)} ground truth JSON files → {args.output_dir}/")
