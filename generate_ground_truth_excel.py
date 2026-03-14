"""
Ground Truth Excel Generator
=============================
Generates a single relational Excel workbook for entering ground truth
values across all 45 forms. Designed for fast manual data entry.

Workbook structure
------------------
Sheet 1 — FORMS (one row per form)
    Columns: form_id, full_name, matric_number, department, faculty,
             level_100, level_200, level_300, level_400,
             session, semester, mode_fulltime, mode_parttime,
             total_units, date
    (student_signature excluded by design)

Sheet 2 — COURSES (one row per course entry, FK to FORMS)
    Columns: form_id, row_number, sn, course_code, course_title, unit

Usage
-----
    python generate_ground_truth_excel.py
    # Opens ground_truth/ground_truth_entry.xlsx
    # Fill in values, save
    # Run: python excel_to_json.py  to convert back to per-form JSON
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Styling constants ──────────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1A1A2E")
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_SUB_FILL   = PatternFill("solid", fgColor="E8EAF6")
_SUB_FONT   = Font(bold=True, color="1A1A2E", size=9)
_NOTE_FILL  = PatternFill("solid", fgColor="FFF9C4")
_BOOL_FILL  = PatternFill("solid", fgColor="E8F5E9")
_THIN       = Side(style="thin", color="CCCCCC")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

N_FORMS = 45


def _header(ws, row: int, col: int, value: str, sub: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font  = _SUB_FONT if sub else _HDR_FONT
    cell.fill  = _SUB_FILL if sub else _HDR_FILL
    cell.alignment = _CENTER
    cell.border    = _BORDER


def _note(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _NOTE_FILL
    cell.font      = Font(italic=True, size=8, color="555555")
    cell.alignment = _CENTER
    cell.border    = _BORDER


def build(output_path: str = "ground_truth/ground_truth_entry.xlsx") -> str:
    wb = openpyxl.Workbook()

    # ── Sheet 1: FORMS ────────────────────────────────────────────────────────
    ws_forms = wb.active
    ws_forms.title = "FORMS"
    ws_forms.freeze_panes = "B3"

    # Title row
    ws_forms.merge_cells("A1:O1")
    title = ws_forms["A1"]
    title.value     = "STUDENT ACADEMIC RECORD — Ground Truth Entry  |  Fill ALL yellow cells  |  Checkboxes: TRUE / FALSE  |  Blank fields: leave empty"
    title.font      = Font(bold=True, size=10, color="1A1A2E")
    title.fill      = _NOTE_FILL
    title.alignment = _CENTER
    ws_forms.row_dimensions[1].height = 28

    # Column definitions: (header_label, col_width, is_checkbox)
    form_cols = [
        ("form_id",       12, False),
        ("full_name",     28, False),
        ("matric_number", 18, False),
        ("department",    22, False),
        ("faculty",       22, False),
        ("level_100",     10, True),
        ("level_200",     10, True),
        ("level_300",     10, True),
        ("level_400",     10, True),
        ("session",       14, False),
        ("semester",      12, False),
        ("mode_fulltime", 12, True),
        ("mode_parttime", 12, True),
        ("total_units",   11, False),
        ("date",          13, False),
    ]

    # Sub-header: group labels
    ws_forms.merge_cells("A2:A2"); _header(ws_forms, 2, 1,  "ID",          sub=True)
    ws_forms.merge_cells("B2:E2"); _header(ws_forms, 2, 2,  "Personal Details", sub=True)
    ws_forms.merge_cells("F2:I2"); _header(ws_forms, 2, 6,  "Level",       sub=True)
    ws_forms.merge_cells("J2:K2"); _header(ws_forms, 2, 10, "Academic",    sub=True)
    ws_forms.merge_cells("L2:M2"); _header(ws_forms, 2, 12, "Mode",        sub=True)
    ws_forms.merge_cells("N2:O2"); _header(ws_forms, 2, 14, "Summary",     sub=True)
    ws_forms.row_dimensions[2].height = 18

    # Column headers (row 3)
    for col_idx, (label, width, _) in enumerate(form_cols, start=1):
        _header(ws_forms, 3, col_idx, label)
        ws_forms.column_dimensions[get_column_letter(col_idx)].width = width
    ws_forms.row_dimensions[3].height = 32

    # Boolean validation for checkbox columns
    bool_dv = DataValidation(
        type="list", formula1='"TRUE,FALSE"',
        allow_blank=True, showDropDown=False
    )
    ws_forms.add_data_validation(bool_dv)

    # Data rows
    for i in range(1, N_FORMS + 1):
        row = i + 3
        ws_forms.row_dimensions[row].height = 18

        for col_idx, (label, _, is_bool) in enumerate(form_cols, start=1):
            cell = ws_forms.cell(row=row, column=col_idx)
            cell.border    = _BORDER
            cell.alignment = _CENTER if is_bool else _LEFT

            if label == "form_id":
                cell.value = f"form_{i:03d}"
                cell.font  = Font(bold=True, size=9)
                cell.fill  = _SUB_FILL
            elif is_bool:
                cell.fill = _BOOL_FILL
                bool_dv.add(cell)
            else:
                cell.fill = PatternFill("solid", fgColor="FFFDE7")
                cell.font = Font(size=9)

    # ── Sheet 2: COURSES ──────────────────────────────────────────────────────
    ws_courses = wb.create_sheet("COURSES")
    ws_courses.freeze_panes = "A3"

    ws_courses.merge_cells("A1:F1")
    t2 = ws_courses["A1"]
    t2.value     = "COURSE REGISTRATION — 3 rows per form  |  Leave row blank if not used  |  form_id links to FORMS sheet"
    t2.font      = Font(bold=True, size=10, color="1A1A2E")
    t2.fill      = _NOTE_FILL
    t2.alignment = _CENTER
    ws_courses.row_dimensions[1].height = 28

    course_cols = [
        ("form_id",      12),
        ("row_number",   10),
        ("sn",           8),
        ("course_code",  14),
        ("course_title", 38),
        ("unit",         8),
    ]
    for col_idx, (label, width) in enumerate(course_cols, start=1):
        _header(ws_courses, 2, col_idx, label)
        ws_courses.column_dimensions[get_column_letter(col_idx)].width = width
    ws_courses.row_dimensions[2].height = 32

    # 3 rows per form = 135 rows
    fill_course = PatternFill("solid", fgColor="FFFDE7")
    fill_id     = PatternFill("solid", fgColor="E8EAF6")
    for i in range(1, N_FORMS + 1):
        for r in range(1, 4):
            row = (i - 1) * 3 + r + 2
            ws_courses.row_dimensions[row].height = 17

            for col_idx, (label, _) in enumerate(course_cols, start=1):
                cell = ws_courses.cell(row=row, column=col_idx)
                cell.border    = _BORDER
                cell.alignment = _CENTER if label in ("row_number", "unit", "sn") else _LEFT
                cell.font      = Font(size=9)

                if label == "form_id":
                    cell.value = f"form_{i:03d}"
                    cell.fill  = fill_id
                    cell.font  = Font(bold=True, size=9)
                elif label == "row_number":
                    cell.value = r
                    cell.fill  = fill_id
                else:
                    cell.fill = fill_course

    # ── Sheet 3: README ───────────────────────────────────────────────────────
    ws_readme = wb.create_sheet("README")
    instructions = [
        ("GROUND TRUTH ENTRY INSTRUCTIONS", True),
        ("", False),
        ("FORMS sheet:", True),
        ("  • Each row = one scanned form", False),
        ("  • form_id matches the filename: form_001.tif → form_001", False),
        ("  • Text fields: type exactly what is written on the form", False),
        ("  • Checkbox fields (level_*, mode_*): type TRUE if ticked, FALSE if not", False),
        ("  • Leave a cell empty if the field is blank on the form", False),
        ("  • date format: DD/MM/YYYY", False),
        ("", False),
        ("COURSES sheet:", True),
        ("  • 3 rows pre-filled per form (matching the 3 table rows on the form)", False),
        ("  • Fill in sn, course_code, course_title, unit for each course row", False),
        ("  • Leave all cells blank for unused rows", False),
        ("  • form_id links this sheet back to the FORMS sheet", False),
        ("", False),
        ("AFTER FILLING IN:", True),
        ("  • Save the file", False),
        ("  • Run:  python excel_to_json.py", False),
        ("  • This converts the spreadsheet into 45 individual JSON ground truth files", False),
        ("  • JSON files land in:  ground_truth/form_001.json ... form_045.json", False),
        ("", False),
        ("EXCLUDED FIELDS (by design):", True),
        ("  • student_signature — excluded from evaluation per project scope", False),
    ]
    ws_readme.column_dimensions["A"].width = 70
    for r_idx, (text, bold) in enumerate(instructions, start=1):
        cell = ws_readme.cell(row=r_idx, column=1, value=text)
        cell.font      = Font(bold=bold, size=10 if bold else 9)
        cell.alignment = _LEFT
        ws_readme.row_dimensions[r_idx].height = 16 if text else 8

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = build("ground_truth/ground_truth_entry.xlsx")
    print(f"Ground truth workbook created: {path}")
    print(f"  Sheet FORMS   : {N_FORMS} rows (one per form)")
    print(f"  Sheet COURSES : {N_FORMS * 3} rows (3 course rows per form)")
    print(f"  Sheet README  : instructions")
    print(f"\nFill in the yellow cells, then run excel_to_json.py")
